import openpyxl
import requests
import os
from io import BytesIO

def generate_test_data_from_excel(excel_path, default_value):
    test_data = []
    try:
        # response = requests.get(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {}
            for index, cell in enumerate(row):
                header = headers[index].lower() if headers[index] else None
                if cell:
                    data[header] = cell
                elif header in default_value:
                    data[header] = default_value[header]
            test_data.append(data)
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
    return test_data

current_directory = os.path.dirname(os.path.abspath(__file__))
excel_filename = "TestData.xlsx"
excel_path = os.path.join(current_directory, excel_filename) 
default_value = {'build': 'default_build', 'state': 'default_state', 'num': 'default_num'}  
test_data = generate_test_data_from_excel(excel_path, default_value)
print(test_data[0]['buid'])
